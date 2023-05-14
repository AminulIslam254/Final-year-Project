import openpyxl
from flask import Flask, request, jsonify

# Load the Excel file
workbook = openpyxl.load_workbook('projectXlsx.xlsx')

app = Flask(__name__)

# Define the route for making changes to the Excel file
@app.route('/change_excel', methods=['POST'])
def change_excel():
    # Get the criteria and new value from the request data
    criteria_column = request.json['criteria_column']
    criteria_value = request.json['criteria_value']
    new_value = request.json['new_value']

    print(criteria_value,new_value)
    # print(workbook.sheetnames)
    # Select the worksheet you want to modify
    worksheet = workbook.active



    # Loop through the rows and make changes based on the criteria
    i=0
    for row in worksheet.iter_rows(min_col=6):
        i+=1
        if i==criteria_value :
            row[0].value = round(new_value,4)
            

    
    # for row in worksheet.iter_rows(min_col=3):
    #     if type(row[0].value) == float:
    #         if row[0].value == float(criteria_value):
    #             row[0].value = new_value
    #     else:
    #         if row[0].value == (criteria_value):
    #             row[0].value = new_value

    # for row in worksheet.iter_rows(min_col=4):
    #     if type(row[0].value) == float:
    #         if row[0].value == float(criteria_value):
    #             row[0].value = new_value
    #     else:
    #         if row[0].value == (criteria_value):
    #             row[0].value = new_value

    # for row in worksheet.iter_rows(min_col=5):
    #     if type(row[0].value) == float:
    #         if row[0].value == float(criteria_value):
    #             row[0].value = new_value
    #     else:
    #         if row[0].value == (criteria_value):
    #             row[0].value = new_value

    # for row in worksheet.iter_rows(min_col=6):
    #     if type(row[0].value) == float:
    #         if row[0].value == criteria_value:
    #             row[0].value = new_value
    #     else:
    #         if row[0].value == (criteria_value):
    #             row[0].value = new_value
        

    
    

    # # Save the changes
    workbook.save('projectXlsx.xlsx')
    workbook.close()

    # Return a success response
    return jsonify({'message': 'Changes saved successfully'})

if __name__ == '__main__':
    app.run(debug=True, port=6063)
